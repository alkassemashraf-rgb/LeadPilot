from typing import List, Any, Optional
import hashlib
import os
import uuid
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select

from sqlmodel import delete as sql_delete

from app.api import deps
from app.core.db import get_db
from app.models.models import Workspace, WorkspaceKnowledgeFile, KnowledgeChunk
from app.schemas.envelope import ResponseEnvelope, wrap_data, wrap_error
from app.core.modules import require_module_enabled, MODULE_KNOWLEDGE_FILES
from app.services.entitlements import require_entitlement
from app.services.knowledge_chunker import create_chunks_for_file

router = APIRouter()

STORAGE_DIR = "storage/knowledge"

# Max extracted text length per file (200k chars)
MAX_EXTRACTED_TEXT = 200_000


@router.post("/files", response_model=ResponseEnvelope[dict], dependencies=[Depends(require_module_enabled(MODULE_KNOWLEDGE_FILES, "write")), Depends(require_entitlement("knowledge_files", increment=True))])
async def upload_knowledge_file(
    file: UploadFile = File(...),
    notes: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
    workspace: Workspace = Depends(deps.get_active_workspace),
) -> Any:
    """Upload a file to the workspace knowledge base."""
    workspace_dir = os.path.join(STORAGE_DIR, str(workspace.id))
    os.makedirs(workspace_dir, exist_ok=True)

    file_id = str(uuid.uuid4())
    extension = os.path.splitext(file.filename or "")[1]
    storage_path = os.path.join(workspace_dir, f"{file_id}{extension}")

    # Save file and compute SHA256
    try:
        file_bytes = await file.read()
        sha256_hash = hashlib.sha256(file_bytes).hexdigest()
        with open(storage_path, "wb") as buffer:
            buffer.write(file_bytes)
    except Exception as e:
        return wrap_error(f"Failed to save file: {str(e)}")

    # SHA256 dedupe check
    existing = await db.execute(
        select(WorkspaceKnowledgeFile).where(
            WorkspaceKnowledgeFile.workspace_id == workspace.id,
            WorkspaceKnowledgeFile.sha256_hash == sha256_hash,
        )
    )
    dupe = existing.scalars().first()
    if dupe:
        os.remove(storage_path)
        return wrap_error(f"Duplicate file: this content already exists as '{dupe.filename}'")

    # Extract text for supported formats
    extracted_text = None
    status = "READY"

    if extension.lower() in [".txt", ".md", ".json", ".csv"]:
        try:
            with open(storage_path, "r", encoding="utf-8") as f:
                extracted_text = f.read()[:MAX_EXTRACTED_TEXT]
        except Exception:
            status = "FAILED"
    elif extension.lower() == ".pdf":
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(storage_path)
            pages_text = [page.get_text() for page in doc]
            doc.close()
            full_text = "\n".join(pages_text)
            if not full_text.strip():
                status = "FAILED"
            else:
                extracted_text = full_text[:MAX_EXTRACTED_TEXT]
        except ImportError:
            status = "FAILED"
        except Exception:
            status = "FAILED"
    elif extension.lower() == ".docx":
        try:
            from docx import Document as DocxDocument
            doc = DocxDocument(storage_path)
            full_text = "\n".join(p.text for p in doc.paragraphs)
            if not full_text.strip():
                status = "FAILED"
            else:
                extracted_text = full_text[:MAX_EXTRACTED_TEXT]
        except ImportError:
            status = "FAILED"
        except Exception:
            status = "FAILED"
    elif extension.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(storage_path, read_only=True, data_only=True)
            rows_text = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    rows_text.append("\t".join(cells))
            wb.close()
            full_text = "\n".join(rows_text)
            if not full_text.strip():
                status = "FAILED"
            else:
                extracted_text = full_text[:MAX_EXTRACTED_TEXT]
        except ImportError:
            status = "FAILED"
        except Exception:
            status = "FAILED"

    knowledge_file = WorkspaceKnowledgeFile(
        id=uuid.UUID(file_id),
        workspace_id=workspace.id,
        filename=file.filename,
        mime_type=file.content_type or "application/octet-stream",
        size_bytes=os.path.getsize(storage_path),
        storage_path=storage_path,
        extracted_text=extracted_text,
        notes=notes,
        sha256_hash=sha256_hash,
        status=status,
    )

    db.add(knowledge_file)
    await db.flush()

    # Create knowledge chunks for retrieval
    chunk_count = 0
    if extracted_text:
        chunk_count = await create_chunks_for_file(
            db, workspace.id, uuid.UUID(file_id), extracted_text,
        )

    await db.commit()
    await db.refresh(knowledge_file)

    return wrap_data({
        "id": str(knowledge_file.id),
        "filename": knowledge_file.filename,
        "extracted": extracted_text is not None,
        "status": status,
        "chunk_count": chunk_count,
    })


@router.get("/files", response_model=ResponseEnvelope[List[dict]], dependencies=[Depends(require_module_enabled(MODULE_KNOWLEDGE_FILES, "read")), Depends(require_entitlement("knowledge_files"))])
async def list_knowledge_files(
    db: AsyncSession = Depends(get_db),
    workspace: Workspace = Depends(deps.get_active_workspace),
) -> Any:
    """List all knowledge files for the active workspace."""
    result = await db.execute(
        select(WorkspaceKnowledgeFile).where(WorkspaceKnowledgeFile.workspace_id == workspace.id)
    )
    files = result.scalars().all()

    return wrap_data([
        {
            "id": str(f.id),
            "filename": f.filename,
            "mime_type": f.mime_type,
            "size_bytes": f.size_bytes,
            "notes": f.notes,
            "status": f.status,
            "extracted": f.extracted_text is not None,
            "created_at": f.created_at.isoformat(),
        }
        for f in files
    ])


@router.get("/files/{file_id}/download", dependencies=[Depends(require_module_enabled(MODULE_KNOWLEDGE_FILES, "read")), Depends(require_entitlement("knowledge_files"))])
async def download_knowledge_file(
    file_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    workspace: Workspace = Depends(deps.get_active_workspace),
) -> FileResponse:
    """Download a knowledge file."""
    result = await db.execute(
        select(WorkspaceKnowledgeFile).where(
            WorkspaceKnowledgeFile.id == file_id,
            WorkspaceKnowledgeFile.workspace_id == workspace.id,
        )
    )
    kb_file = result.scalars().first()
    if not kb_file:
        raise HTTPException(status_code=404, detail="File not found")
    if not os.path.exists(kb_file.storage_path):
        raise HTTPException(status_code=404, detail="File missing from storage")

    return FileResponse(
        kb_file.storage_path,
        filename=kb_file.filename,
        media_type=kb_file.mime_type,
    )


@router.delete("/files/{file_id}", response_model=ResponseEnvelope[dict], dependencies=[Depends(require_module_enabled(MODULE_KNOWLEDGE_FILES, "write")), Depends(require_entitlement("knowledge_files"))])
async def delete_knowledge_file(
    file_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    workspace: Workspace = Depends(deps.get_active_workspace),
) -> Any:
    """Delete a knowledge file."""
    result = await db.execute(
        select(WorkspaceKnowledgeFile).where(
            WorkspaceKnowledgeFile.id == file_id,
            WorkspaceKnowledgeFile.workspace_id == workspace.id,
        )
    )
    kb_file = result.scalars().first()
    if not kb_file:
        raise HTTPException(status_code=404, detail="File not found")

    # Delete chunks first
    await db.execute(
        sql_delete(KnowledgeChunk).where(KnowledgeChunk.knowledge_file_id == file_id)
    )

    # Delete from storage
    if os.path.exists(kb_file.storage_path):
        os.remove(kb_file.storage_path)

    await db.delete(kb_file)
    await db.commit()

    return wrap_data({"deleted": True})
